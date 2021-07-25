from openpyxl import *
from tkinter import ttk, IntVar, messagebox, Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename


class Compare:
    def __init__(self, file_adrs, column_a, column_b, barcode):
        self.file = file_adrs
        self.a = column_a
        self.b = column_b
        self.barcode = barcode
    
    def compare(self):
        file = load_workbook(self.file)

        worksheet = file.active
        a = worksheet[self.a]
        b = worksheet[self.b]
        brcode = worksheet[self.barcode]

        columns = {}
        br_list = []
        count = 0

        for cell in brcode:
                columns[cell.value] = (a[count].value, b[count].value)
                count += 1

        for br in brcode:
            br_list.append(br.value)


        self.increse_cost = {}
        self.decrese_cost = {}
        self.const_cost = {}

        for code in br_list:
            costs = columns.get(code)
            try:
                if int(costs[0]) > int(costs[1]):
                    self.increse_cost[code] = costs

                elif int(costs[0]) < int(costs[1]):
                    self.decrese_cost[code] = costs

                else:
                    self.const_cost[code] = costs

            except ValueError:
                pass


    def save_increse(self, file):
        increse_cost = self.increse_cost

        workbook = Workbook()
        worksheet = workbook.active

        row = 1
        for code in increse_cost:
            worksheet[f"A{row}"] = code

            cost = increse_cost.get(code)
            worksheet[f"B{row}"] = cost[0]
            worksheet[f"C{row}"] = cost[1]

            row += 1

            
        workbook.save(filename=f"{file}")

    def save_decrese(self, file):

        decrese_cost = self.decrese_cost

        workbook = Workbook()
        worksheet = workbook.active

        row = 1
        for code in decrese_cost:
            worksheet[f"A{row}"] = code

            cost = decrese_cost.get(code)
            worksheet[f"B{row}"] = cost[0]
            worksheet[f"C{row}"] = cost[1]

            row += 1

            
        workbook.save(filename=f"{file}")

    def save_const(self, file):
        const_cost = self.const_cost

        workbook = Workbook()
        worksheet = workbook.active

        row = 1
        for code in const_cost:
            worksheet[f"A{row}"] = str(code)

            cost = const_cost.get(code)
            worksheet[f"B{row}"] = str(cost[0])
            worksheet[f"C{row}"] = str(cost[1])

            row += 1

            
        workbook.save(filename=f"{file}")

class CostGUI:
    def __init__(self, master):
        self.master = master
        self.var = IntVar()

        ttk.Label(self.master, text="ستون ها را برای مقایسه انتخاب کنید").place(x=10, y=10)

        ttk.Label(self.master, text="ستون اول").place(x=10, y=50)
        self.column_one = ttk.Entry(self.master, width=5)
        self.column_one.place(x=100, y=50)

        ttk.Label(self.master, text="ستون دوم").place(x=10, y=100)
        self.column_two = ttk.Entry(self.master, width=5)
        self.column_two.place(x=100, y=100)

        ttk.Label(self.master, text="انتخاب ستون بارکد").place(x=10, y=150)
        self.barcode = ttk.Entry(self.master, width=5)
        self.barcode.place(x=200, y=150)

        ttk.Radiobutton(self.master, text="گزارش افزایش قیمت", value=1, variable=self.var).place(x=10, y=200)
        ttk.Radiobutton(self.master, text="گزارش کاهش قیمت", value=2, variable=self.var).place(x=10, y=250)
        ttk.Radiobutton(self.master, text="گزارش ثابت قیمت", value=3, variable=self.var).place(x=10, y=300)

        ttk.Button(self.master, text="انجلم عملیات", command=self.do_func).place(x=10, y=350)

    def do_func(self):

        column_one = self.column_one.get().upper()
        column_two = self.column_two.get().upper()
        barcode = self.barcode.get().upper()

        file_dir = askopenfilename(title="فایل را انتخاب کنید ...", filetypes=(("Excell", "*.xlsx"),("All", "*.*")))

        cmpr = Compare(file_dir, column_one, column_two, barcode)
        cmpr.compare()

        if self.var.get() == 1:
            save_dir = asksaveasfilename(title="ذخیره فایل ...", filetypes=(("Excell", "*.xlsx"),("All", "*.*")))
            cmpr.save_increse(save_dir + ".xlsx")
            messagebox.showinfo("", "عملیات انجام شد")

        elif self.var.get() == 2:
            save_dir = asksaveasfilename(title="ذخیره فایل ...", filetypes=(("Excell", "*.xlsx"),("All", "*.*")))
            cmpr.save_decrese(save_dir + ".xlsx")
            messagebox.showinfo("", "عملیات انجام شد")

        else:
            save_dir = asksaveasfilename(title="ذخیره فایل ...", filetypes=(("Excell", "*.xlsx"),("All", "*.*")))
            cmpr.save_const(save_dir + ".xlsx")
            messagebox.showinfo("", "عملیات انجام شد")

