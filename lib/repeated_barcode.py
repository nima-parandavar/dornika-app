from openpyxl import *
from tkinter import messagebox, ttk, IntVar
from tkinter.filedialog import askopenfilename, asksaveasfilename


class RepeatedQr:
    def __init__(self, main_file, second_file):

        self.main_file = main_file
        self.sec_file = second_file

    def load_columns(self, main, second, m1, m2, m3, m4, m5, m6, m7, m8, m9, m10, s1, s2, s3, s4, s5, s6, s7, s8, s9, s10):
        self.main_columns = {
            "main": main.upper(),
            "m1": m1.upper(),
            "m2": m2.upper(),
            "m3": m3.upper(),
            "m4": m4.upper(),
            "m5": m5.upper(),
            "m6": m6.upper(),
            "m5": m7.upper(),
            "m5": m8.upper(),
            "m5": m9.upper(),
            "m5": m10.upper(),
        }

        self.sec_columns = {
            "sec": second.upper(),
            "s1": s1.upper(),
            "s2": s2.upper(),
            "s3": s3.upper(),
            "s4": s4.upper(),
            "s5": s5.upper(),
            "s6": s6.upper(),
            "s7": s7.upper(),
            "s8": s8.upper(),
            "s9": s9.upper(),
            "s10": s10.upper(),
        }

    def find_similer(self, file_dir, state):
        # Columns from Main File
        main_columns = self.main_columns
        main = main_columns.get("main")
        m1 = main_columns.get("m1")
        m2 = main_columns.get("m2")
        m3 = main_columns.get("m3")
        m4 = main_columns.get("m4")
        m5 = main_columns.get("m5")
        m6 = main_columns.get("m6")
        m7 = main_columns.get("m7")
        m8 = main_columns.get("m8")
        m9 = main_columns.get("m9")
        m10 = main_columns.get("m10")

        # Columns from second File
        sec_columns = self.sec_columns
        sec = sec_columns.get("sec")
        s1 = sec_columns.get("s1")
        s2 = sec_columns.get("s2")
        s3 = sec_columns.get("s3")
        s4 = sec_columns.get("s4")
        s5 = sec_columns.get("s5")
        s6 = sec_columns.get("s6")
        s7 = sec_columns.get("s7")
        s8 = sec_columns.get("s8")
        s9 = sec_columns.get("s9")
        s10 = sec_columns.get("s10")

        # Load files
        main_file = self.main_file
        sec_file = self.sec_file
        # parent file
        pwb = load_workbook(main_file)
        pws = pwb.active
        pw = pws[main]
        # Childe File
        cwb = load_workbook(sec_file)
        cws = cwb.active
        cw = cws[sec]

        main_dic = {}
        for parent_qr in range(len(pw)):
            main_dic[str(pw[parent_qr].value)] = parent_qr + 1

        sec_code_list = []
        for child_qr in range(len(cw)):
            sec_code_list.append(str(cw[child_qr].value))

        # Load new excell
        workbook = Workbook()
        worksheet = workbook.active

        row = 1
        sf_row = 1

        if state == 1:
            for code in sec_code_list:
                val = main_dic.get(code)
                if val:
                    worksheet[f"A{row}"] = code
                    worksheet[f"B{row}"] = pws[f"{m1}{val}"].value
                    worksheet[f"C{row}"] = pws[f"{m2}{val}"].value
                    worksheet[f"D{row}"] = pws[f"{m3}{val}"].value
                    worksheet[f"E{row}"] = pws[f"{m4}{val}"].value
                    worksheet[f"F{row}"] = pws[f"{m5}{val}"].value
                    worksheet[f"G{row}"] = pws[f"{m6}{val}"].value
                    worksheet[f"H{row}"] = pws[f"{m7}{val}"].value
                    worksheet[f"I{row}"] = pws[f"{m8}{val}"].value
                    worksheet[f"J{row}"] = pws[f"{m9}{val}"].value
                    worksheet[f"K{row}"] = pws[f"{m10}{val}"].value

                    row += 1
            # save file
            workbook.save(filename=f"{file_dir}.xlsx")
        elif state == 2:
            for code in sec_code_list:
                val = main_dic.get(code)
                if val:
                    worksheet[f"A{row}"] = code
                    worksheet[f"B{row}"] = cws[f"{s1}{sf_row}"].value
                    worksheet[f"C{row}"] = cws[f"{s2}{sf_row}"].value
                    worksheet[f"D{row}"] = cws[f"{s3}{sf_row}"].value
                    worksheet[f"E{row}"] = cws[f"{s4}{sf_row}"].value
                    worksheet[f"F{row}"] = cws[f"{s5}{sf_row}"].value
                    worksheet[f"G{row}"] = cws[f"{s6}{sf_row}"].value
                    worksheet[f"H{row}"] = cws[f"{s7}{sf_row}"].value
                    worksheet[f"I{row}"] = cws[f"{s8}{sf_row}"].value
                    worksheet[f"J{row}"] = cws[f"{s9}{sf_row}"].value
                    worksheet[f"K{row}"] = cws[f"{s10}{sf_row}"].value

                    row += 1
                sf_row += 1
            # save file
            workbook.save(filename=f"{file_dir}.xlsx")
        elif state == 3:
            for code in sec_code_list:
                val = main_dic.get(code)
                if val:
                    worksheet[f"A{row}"] = code
                    worksheet[f"B{row}"] = pws[f"{m1}{val}"].value
                    worksheet[f"C{row}"] = pws[f"{m2}{val}"].value
                    worksheet[f"D{row}"] = pws[f"{m3}{val}"].value
                    worksheet[f"E{row}"] = pws[f"{m4}{val}"].value
                    worksheet[f"F{row}"] = pws[f"{m5}{val}"].value
                    worksheet[f"G{row}"] = pws[f"{m6}{val}"].value
                    worksheet[f"H{row}"] = pws[f"{m7}{val}"].value
                    worksheet[f"I{row}"] = pws[f"{m8}{val}"].value
                    worksheet[f"J{row}"] = pws[f"{m9}{val}"].value
                    worksheet[f"K{row}"] = pws[f"{m10}{val}"].value

                    worksheet[f"L{row}"] = cws[f"{s1}{sf_row}"].value
                    worksheet[f"M{row}"] = cws[f"{s2}{sf_row}"].value
                    worksheet[f"N{row}"] = cws[f"{s3}{sf_row}"].value
                    worksheet[f"O{row}"] = cws[f"{s4}{sf_row}"].value
                    worksheet[f"P{row}"] = cws[f"{s5}{sf_row}"].value
                    worksheet[f"Q{row}"] = cws[f"{s6}{sf_row}"].value
                    worksheet[f"R{row}"] = cws[f"{s7}{sf_row}"].value
                    worksheet[f"S{row}"] = cws[f"{s8}{sf_row}"].value
                    worksheet[f"T{row}"] = cws[f"{s9}{sf_row}"].value
                    worksheet[f"U{row}"] = cws[f"{s10}{sf_row}"].value

                    row += 1
                sf_row += 1
            # save file
            workbook.save(filename=f"{file_dir}.xlsx")
        else:
            messagebox.showerror("خطا", "بین گزینه ها باید یکی را انتخاب کنید")
            return

    def find_unsimiler(self, file_dir):

        # Columns from Main File
        main_columns = self.main_columns
        main = main_columns.get("main")

        # Columns from second File
        sec_columns = self.sec_columns

        sec = sec_columns.get("sec")
        s1 = sec_columns.get("s1")
        s2 = sec_columns.get("s2")
        s3 = sec_columns.get("s3")
        s4 = sec_columns.get("s4")
        s5 = sec_columns.get("s5")
        s6 = sec_columns.get("s6")
        s7 = sec_columns.get("s7")
        s8 = sec_columns.get("s8")
        s9 = sec_columns.get("s9")
        s10 = sec_columns.get("s10")

        # Load files
        main_file = self.main_file
        sec_file = self.sec_file
        # parent file
        pwb = load_workbook(main_file)
        pws = pwb.active
        pw = pws[main]
        # Childe File
        cwb = load_workbook(sec_file)
        cws = cwb.active
        cw = cws[sec]

        main_dic = {}
        for parent_qr in range(len(pw)):
            main_dic[str(pw[parent_qr].value)] = parent_qr + 1

        sec_code_list = []
        for child_qr in range(len(cw)):
            sec_code_list.append(str(cw[child_qr].value))

        workbook = Workbook()
        worksheet = workbook.active

        row = 1
        sf_row = 1

        for code in sec_code_list:
            val = main_dic.get(code)
            if not val:
                worksheet[f"A{row}"] = code
                worksheet[f"B{row}"] = cws[f"{s1}{sf_row}"].value
                worksheet[f"C{row}"] = cws[f"{s2}{sf_row}"].value
                worksheet[f"D{row}"] = cws[f"{s3}{sf_row}"].value
                worksheet[f"E{row}"] = cws[f"{s4}{sf_row}"].value
                worksheet[f"F{row}"] = cws[f"{s5}{sf_row}"].value
                worksheet[f"G{row}"] = cws[f"{s6}{sf_row}"].value
                worksheet[f"H{row}"] = cws[f"{s7}{sf_row}"].value
                worksheet[f"I{row}"] = cws[f"{s8}{sf_row}"].value
                worksheet[f"G{row}"] = cws[f"{s9}{sf_row}"].value
                worksheet[f"K{row}"] = cws[f"{s10}{sf_row}"].value

                row += 1
            sf_row += 1

        # save file
        workbook.save(filename=f"{file_dir}.xlsx")


class GUI:
    def __init__(self):
        self.int_var = IntVar()
        self.state = IntVar()

        self.main_file_adrs: str
        self.second_file_adrs: str

    def set_column(self, master):
        ttk.Label(master, text="ستون های فایل اصلی و فایل ثانویه برای مقایسه مشخص کنید").place(
            x=10, y=10)

        ttk.Label(master, text="بارکد فایل اصلی").place(x=10, y=50)
        self.main = ttk.Entry(master, width=3)
        self.main.place(x=150, y=50)

        ttk.Label(master, text="بارکد فایل ثانویه").place(x=10, y=100)
        self.sec = ttk.Entry(master, width=3)
        self.sec.place(x=150, y=100)

        ttk.Label(master, text="دیگر ستون های فایل اصلی را مشخص کنید (در صورت موجود)").place(
            x=10, y=150)

        ttk.Label(master, text="ستون ۱").place(x=10, y=200)
        self.m1 = ttk.Entry(master, width=3)
        self.m1.place(x=10, y=250)

        ttk.Label(master, text="ستون ۲").place(x=100, y=200)
        self.m2 = ttk.Entry(master, width=3)
        self.m2.place(x=100, y=250)

        ttk.Label(master, text="ستون ۳").place(x=200, y=200)
        self.m3 = ttk.Entry(master, width=3)
        self.m3.place(x=200, y=250)

        ttk.Label(master, text="ستون ۴").place(x=300, y=200)
        self.m4 = ttk.Entry(master, width=3)
        self.m4.place(x=300, y=250)

        ttk.Label(master, text="ستون ۵").place(x=400, y=200)
        self.m5 = ttk.Entry(master, width=3)
        self.m5.place(x=400, y=250)

        ttk.Label(master, text="ستون ۶").place(x=500, y=200)
        self.m6 = ttk.Entry(master, width=3)
        self.m6.place(x=500, y=250)

        ttk.Label(master, text="ستون ۷").place(x=600, y=200)
        self.m7 = ttk.Entry(master, width=3)
        self.m7.place(x=600, y=250)

        ttk.Label(master, text="ستون ۸").place(x=700, y=200)
        self.m8 = ttk.Entry(master, width=3)
        self.m8.place(x=700, y=250)

        ttk.Label(master, text="ستون ۹").place(x=800, y=200)
        self.m9 = ttk.Entry(master, width=3)
        self.m9.place(x=800, y=250)

        ttk.Label(master, text="ستون ۱۰").place(x=900, y=200)
        self.m10 = ttk.Entry(master, width=3)
        self.m10.place(x=900, y=250)

        ttk.Label(master, text="دیگر ستون های فایل ثانویه را مشخص کنید (در صورت موجود)").place(
            x=10, y=300)

        ttk.Label(master, text="ستون ۱").place(x=10, y=350)
        self.s1 = ttk.Entry(master, width=3)
        self.s1.place(x=10, y=400)

        ttk.Label(master, text="ستون ۲").place(x=100, y=350)
        self.s2 = ttk.Entry(master, width=3)
        self.s2.place(x=100, y=400)

        ttk.Label(master, text="ستون ۳").place(x=200, y=350)
        self.s3 = ttk.Entry(master, width=3)
        self.s3.place(x=200, y=400)

        ttk.Label(master, text="ستون ۴").place(x=300, y=350)
        self.s4 = ttk.Entry(master, width=3)
        self.s4.place(x=300, y=400)

        ttk.Label(master, text="ستون ۵").place(x=400, y=350)
        self.s5 = ttk.Entry(master, width=3)
        self.s5.place(x=400, y=400)

        ttk.Label(master, text="ستون ۶").place(x=500, y=350)
        self.s6 = ttk.Entry(master, width=3)
        self.s6.place(x=500, y=400)

        ttk.Label(master, text="ستون ۷").place(x=600, y=350)
        self.s7 = ttk.Entry(master, width=3)
        self.s7.place(x=600, y=400)

        ttk.Label(master, text="ستون ۸").place(x=700, y=350)
        self.s8 = ttk.Entry(master, width=3)
        self.s8.place(x=700, y=400)

        ttk.Label(master, text="ستون ۹").place(x=800, y=350)
        self.s9 = ttk.Entry(master, width=3)
        self.s9.place(x=800, y=400)

        ttk.Label(master, text="ستون ۱۰").place(x=900, y=350)
        self.s10 = ttk.Entry(master, width=3)
        self.s10.place(x=900, y=400)

    def load_files(self, master):
        ttk.Label(master, text="انتخاب فایل اصلی").place(x=10, y=10)

        mfal = ttk.Label(master)
        mfal.place(x=200, y=50)

        def open_main_file():
            main_file = askopenfilename(title="انتخاب ...", filetypes=(
                ("Excel", "*.xlsx"), ("All", "*.*")))
            mfal.config(text=main_file)
            self.main_file_adrs = main_file

        ttk.Button(master, text="انتخاب",
                   command=open_main_file).place(x=10, y=50)

        ttk.Label(master, text="انتخاب فایل ثانویه").place(x=10, y=100)
        sfal = ttk.Label(master)
        sfal.place(x=200, y=150)

        def choose_sec_file():
            sec_file = askopenfilename(title="انتخاب ...", filetypes=(
                ("Excel", "*.xlsx"), ("All", "*.*")))
            sfal.config(text=sec_file)
            self.second_file_adrs = sec_file

        ttk.Button(master, text="انتخاب",
                   command=choose_sec_file).place(x=10, y=150)

    def seprate_file_gui(self, master):
        ttk.Label(master, text="نوع عملیات را مشخص کنید").place(x=10, y=10)

        ttk.Radiobutton(master, text="فقط فایل های مشابه",
                        value=1, variable=self.int_var).place(x=10, y=50)
        ttk.Radiobutton(master, text="فقط فایل های غیر مشابه",
                        value=2, variable=self.int_var).place(x=10, y=100)

        ttk.Label(master, text="انتخاب ستون های مورد نیاز").place(x=10, y=150)

        ttk.Radiobutton(master, text="ستون های فایل اصلی",
                        value=1, variable=self.state).place(x=10, y=200)
        ttk.Radiobutton(master, text="ستون های فایل قانویه",
                        value=2, variable=self.state).place(x=10, y=230)
        ttk.Radiobutton(master, text="ستون های هردو فایل را می خواهم",
                        value=3, variable=self.state).place(x=10, y=260)

        ttk.Button(master, text="انجام عملیات",
                   command=self.do_func).place(x=10, y=300)

    def do_func(self):
        try:
            rqr = RepeatedQr(self.main_file_adrs, self.second_file_adrs)

            rqr.load_columns(
                self.main.get(),
                self.sec.get(),
                self.m1.get() if self.m1.get() != "" else "Z",
                self.m2.get() if self.m2.get() != "" else "Z",
                self.m3.get() if self.m3.get() != "" else "Z",
                self.m4.get() if self.m4.get() != "" else "Z",
                self.m5.get() if self.m5.get() != "" else "Z",
                self.m6.get() if self.m6.get() != "" else "Z",
                self.m7.get() if self.m7.get() != "" else "Z",
                self.m8.get() if self.m8.get() != "" else "Z",
                self.m9.get() if self.m9.get() != "" else "Z",
                self.m10.get() if self.m10.get() != "" else "Z",
                self.s1.get() if self.s1.get() != "" else "Z",
                self.s2.get() if self.s2.get() != "" else "Z",
                self.s3.get() if self.s3.get() != "" else "Z",
                self.s4.get() if self.s4.get() != "" else "Z",
                self.s5.get() if self.s5.get() != "" else "Z",
                self.s6.get() if self.s6.get() != "" else "Z",
                self.s7.get() if self.s7.get() != "" else "Z",
                self.s8.get() if self.s8.get() != "" else "Z",
                self.s9.get() if self.s9.get() != "" else "Z",
                self.s10.get() if self.s10.get() != "" else "Z",
            )

            if self.int_var.get() == 1:

                save_file_adrs = asksaveasfilename(
                    title="ذخیره در ...", filetypes=(("Excell", "*.xlsx"), ("All", "*.*")))
                rqr.find_similer(save_file_adrs, self.state.get())
                messagebox.showinfo(
                    "عملیات انجام شد", "فایل \n {} \n ذخیره شد".format(save_file_adrs))

            elif self.int_var.get() == 2:

                save_file_adrs = asksaveasfilename(
                    title="ذخیره در ...", filetypes=(("Excell", "*.xlsx"), ("All", "*.*")))
                rqr.find_unsimiler(save_file_adrs)
                messagebox.showinfo(
                    "عملیات انجام شد", "فایل \n {} \n ذخیره شد".format(save_file_adrs))

            else:

                messagebox.showerror(
                    "خطا", "بین گزینه مشابه و غیر مشابه یکی را انتخاب کنید")

        except AttributeError:
            messagebox.showerror("خطا", "فایل ها را انتخاب کنید")
